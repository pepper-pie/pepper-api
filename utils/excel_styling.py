from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd


def apply_table_styles(
    worksheet: Worksheet, 
    df: pd.DataFrame, 
    startrow: int = 0, 
    startcol: int = 0, 
    use_alternate_row_color: bool = True
) -> None:
    """
    Apply table styles to a specific section in the Excel worksheet.

    Args:
        worksheet (Worksheet): The worksheet object to style.
        df (pd.DataFrame): The DataFrame used to calculate dimensions and apply styling.
        startrow (int): The starting row for styling.
        startcol (int): The starting column for styling.
        use_alternate_row_color (bool): Whether to apply alternate row coloring.
    """
    # Header Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_border = Border(
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )

    for col_num, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=startrow + 1, column=startcol + col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Alternate Row Colors (conditionally applied)
    light_gray_fill = PatternFill(start_color="D7D7D7", end_color="D7D7D7", fill_type="solid")
    for row in range(startrow + 2, startrow + len(df) + 2):  # Skip header row
        row_fill = light_gray_fill if use_alternate_row_color and row % 2 == 0 else None
        for col_num in range(startcol + 1, startcol + len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col_num)
            if row_fill:
                cell.fill = row_fill


def format_currency_columns(
    worksheet: Worksheet, 
    df: pd.DataFrame, 
    currency_columns: list, 
    startrow: int = 0, 
    startcol: int = 0
) -> None:
    """
    Format specified columns as currency in a specific section.

    Args:
        worksheet (Worksheet): The worksheet object to style.
        df (pd.DataFrame): The DataFrame containing data.
        currency_columns (list): List of column names to format as currency.
        startrow (int): The starting row for formatting.
        startcol (int): The starting column for formatting.
    """
    currency_format = '₹ #,##0.00'  # INR currency format
    for row in range(startrow + 2, startrow + len(df) + 2):  # Skip header row
        for col_num, column in enumerate(df.columns, start=1):
            if column in currency_columns:
                cell = worksheet.cell(row=row, column=startcol + col_num)
                if isinstance(df[column].iloc[row - startrow - 2], (int, float)):
                    cell.value = float(df[column].iloc[row - startrow - 2])  # Ensure numeric
                cell.number_format = currency_format


def apply_bottom_border(worksheet: Worksheet, df: pd.DataFrame, startrow: int = 0, startcol: int = 0) -> None:
    """
    Apply a thick bottom border to the last row of a specific section.

    Args:
        worksheet (Worksheet): The worksheet object to style.
        df (pd.DataFrame): The DataFrame used to calculate dimensions.
        startrow (int): The starting row for the section.
        startcol (int): The starting column for the section.
    """
    bottom_row = startrow + len(df) + 1
    bottom_border = Border(
        bottom=Side(border_style="medium", color="000000"),
    )
    for col_num in range(startcol + 1, startcol + len(df.columns) + 1):
        cell = worksheet.cell(row=bottom_row, column=col_num)
        cell.border = bottom_border
        
        

def sentence_case(name: str) -> str:
    return name.capitalize().replace('_', ' ').replace('-', ' ')